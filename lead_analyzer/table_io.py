"""Tabellen-Ein-/Ausgabe: xlsx (openpyxl) + csv (stdlib).

Liest die Eingabe unter Erhalt von Spaltenreihenfolge und Originalwerten ein,
erkennt die URL-Spalte tolerant und schreibt die Ausgabe = alle Original-Spalten
unverändert + genau zwei Score-Spalten (+ optionale Begründungsspalte).

Kein pandas: pandas würde Spaltentypen still umcasten und damit AC2 «unverändert»
verletzen.
"""

from __future__ import annotations

import csv
import json
import os
from typing import Iterable

from openpyxl import Workbook, load_workbook

from .models import RowRecord, RowResult

# Exakte Header der zwei Score-Spalten (CLAUDE.md §3).
COL_BEDARF = "Website-Bedarf (1-5)"
COL_ZAHL = "Zahlungskräftigkeit (1-5)"
COL_REASON = "Begründung"

# Tolerante Erkennung der URL-Spalte: Header (lowercase, getrimmt) gegen diese Marker.
_URL_HINTS = ("url", "website", "webseite", "web", "homepage", "internet", "domain", "link", "site")


class InputError(Exception):
    """Eingabe ist unbrauchbar (z.B. keine erkennbare URL-Spalte) — klare Meldung statt Absturz."""


def read_rows(path: str) -> tuple[list[str], list[RowRecord]]:
    """Liest Eingabedatei -> (Header-Liste, Zeilen). Erhält Spaltenreihenfolge.

    Unterstützt .xlsx und .csv (Endung entscheidet). Leere Header werden zu
    'Spalte N' aufgefüllt, Duplikate eindeutig gemacht, damit nichts kollidiert.
    """
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        headers, rows = _read_xlsx(path)
    elif ext in (".csv", ".tsv", ".txt"):
        headers, rows = _read_csv(path, delimiter="\t" if ext == ".tsv" else ",")
    else:
        raise InputError(
            f"Nicht unterstütztes Eingabeformat '{ext}'. Erlaubt: .xlsx, .csv"
        )
    if not headers:
        raise InputError("Eingabedatei hat keine Kopfzeile / ist leer.")
    return headers, rows


def _dedupe_headers(raw: list[object]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for i, h in enumerate(raw):
        name = ("" if h is None else str(h)).strip()
        if not name:
            name = f"Spalte {i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name} ({seen[name]})"
        else:
            seen[name] = 1
        headers.append(name)
    return headers


def _read_xlsx(path: str) -> tuple[list[str], list[RowRecord]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], []
    headers = _dedupe_headers(list(header_row))
    records: list[RowRecord] = []
    idx = 0
    for values in rows_iter:
        values = list(values)
        # Komplett leere Zeilen überspringen (typische Excel-Trailing-Rows).
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
            continue
        cells = {}
        for j, name in enumerate(headers):
            cells[name] = values[j] if j < len(values) else None
        records.append(RowRecord(index=idx, cells=cells))
        idx += 1
    wb.close()
    return headers, records


def _read_csv(path: str, delimiter: str = ",") -> tuple[list[str], list[RowRecord]]:
    # utf-8-sig entfernt ein evtl. BOM; errors=replace verhindert Decode-Abstürze.
    with open(path, "r", encoding="utf-8-sig", newline="", errors="replace") as fh:
        reader = csv.reader(fh, delimiter=delimiter)
        try:
            header_row = next(reader)
        except StopIteration:
            return [], []
        headers = _dedupe_headers(header_row)
        records: list[RowRecord] = []
        idx = 0
        for values in reader:
            if not any((v or "").strip() for v in values):
                continue
            cells = {name: (values[j] if j < len(values) else None) for j, name in enumerate(headers)}
            records.append(RowRecord(index=idx, cells=cells))
            idx += 1
    return headers, records


def detect_url_column(headers: list[str]) -> str:
    """Erkennt die URL-Spalte tolerant. Wirft InputError, wenn keine gefunden wird (AC2/AC4)."""
    lowered = [(h, h.strip().lower()) for h in headers]
    # 1. exakte/sehr nahe Treffer zuerst
    for original, low in lowered:
        if low in _URL_HINTS:
            return original
    # 2. Substring-Treffer (z.B. "Website URL", "Web-Adresse")
    for original, low in lowered:
        if any(hint in low for hint in _URL_HINTS):
            return original
    raise InputError(
        "Keine URL-Spalte erkannt. Erwartet einen Spaltennamen wie "
        f"'URL', 'Website', 'Webseite', 'Web' o.ä. Gefundene Spalten: {headers}"
    )


def _output_headers(headers: list[str], reason_column: bool) -> list[str]:
    out = list(headers) + [COL_BEDARF, COL_ZAHL]
    if reason_column:
        out.append(COL_REASON)
    return out


def write_output(
    path: str,
    headers: list[str],
    ordered: Iterable[tuple[RowRecord, RowResult]],
    reason_column: bool = True,
    write_csv: bool = False,
) -> None:
    """Schreibt die Ausgabe. `ordered` = (RowRecord, RowResult) bereits sortiert.

    Original-Zellen werden 1:1 in Original-Spaltenreihenfolge geschrieben, danach
    die zwei (bzw. drei) neuen Spalten. Score-Werte als echte Integer (numerisch
    sortier-/filterbar in Excel).
    """
    ordered = list(ordered)
    out_headers = _output_headers(headers, reason_column)
    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)

    if path.lower().endswith(".csv") or write_csv:
        csv_path = path if path.lower().endswith(".csv") else os.path.splitext(path)[0] + ".csv"
        _write_csv(csv_path, headers, out_headers, ordered, reason_column)
        if not path.lower().endswith(".csv"):
            _write_xlsx(path, headers, out_headers, ordered, reason_column)
    else:
        _write_xlsx(path, headers, out_headers, ordered, reason_column)


def _row_values(record: RowRecord, result: RowResult, headers: list[str], reason_column: bool) -> list[object]:
    values: list[object] = [record.cells.get(h) for h in headers]
    values.append(int(result.bedarf))
    values.append(int(result.zahl))
    if reason_column:
        values.append(result.reason or "")
    return values


def _write_xlsx(path, headers, out_headers, ordered, reason_column) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(out_headers)
    for record, result in ordered:
        ws.append(_row_values(record, result, headers, reason_column))
    wb.save(path)


def _write_csv(path, headers, out_headers, ordered, reason_column) -> None:
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(out_headers)
        for record, result in ordered:
            writer.writerow(_row_values(record, result, headers, reason_column))


def run_log_path(output: str) -> str:
    """Leitet den Lauf-Log-Pfad aus dem Ausgabepfad ab (gleicher Ordner/Stamm).

    z.B. output/leads_scored.xlsx -> output/leads_scored.run.jsonl
    """
    base, _ = os.path.splitext(output)
    return base + ".run.jsonl"


def write_run_log(path: str, url_col: str, ordered: Iterable[tuple[RowRecord, RowResult]]) -> None:
    """Schreibt ein JSONL-Lauf-Log — IMMER, unabhängig von der Begründungsspalte.

    Ein Objekt pro Zeile mit beiden Scores plus den treibenden Signalen (Bedarf-
    Dimensionen + Zahlungskräftigkeit-Annahmen). So bleibt die Nachvollziehbarkeit
    (AC6) erhalten, selbst wenn die Excel-Begründungsspalte via --no-reason fehlt.
    Eine Zeile mit untypisierbaren Werten darf das Log nicht killen (default=str, AC4).
    """
    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        for record, result in ordered:
            entry = {
                "index": record.index,
                "url": record.cells.get(url_col),
                "bedarf": int(result.bedarf),
                "zahl": int(result.zahl),
                "reason": result.reason or "",
                "bedarf_signals": [
                    {"dim": v.dim, "level": v.level, "reason": v.reason,
                     "source": v.source, "dead": v.dead}
                    for v in result.verdicts
                ],
                "zahl_signals": list(result.zahl_signals),
            }
            fh.write(json.dumps(entry, ensure_ascii=False, default=str) + "\n")
