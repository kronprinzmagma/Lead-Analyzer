"""Wave-0 RED tests for the second-sheet structure in table_io.

These tests MUST FAIL before the second sheet is implemented (RED phase).
They serve as the specification for Task 3 (GREEN implementation).

Requirements covered:
- DIFF-04: output xlsx has 2 sheets; sheet 2 = "Verkaufsargumente" with 3 headers
           + one row per company in the same sorted order as "Leads"
- DIFF-04: "Leads" sheet unchanged (regression guard)
- DIFF-04: companion *_argumente.csv written alongside the main CSV
"""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl
import pytest

from lead_analyzer.config import Config
from lead_analyzer.pipeline import run


# ---------- helper (mirrors test_phase1_io._make_xlsx) ----------

def _make_xlsx(path, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(path)


# ---------- test_second_sheet_structure (DIFF-04) ----------

def test_second_sheet_structure(tmp_path):
    """Output xlsx has 2 sheets; sheet 2 has 3 correct headers + 1 row per company."""
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    headers = ["Kundenname", "Website"]
    rows = [
        ["Firma A", "https://a.ch"],
        ["Firma B", "https://b.ch"],
    ]
    _make_xlsx(inp, headers, rows)

    run(Config(input=str(inp), output=str(out)))

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Leads", "Verkaufsargumente"], (
        f"Expected ['Leads', 'Verkaufsargumente'], got {wb.sheetnames}"
    )

    arg_ws = wb["Verkaufsargumente"]
    all_rows = list(arg_ws.iter_rows(values_only=True))

    # Header row check
    assert list(all_rows[0]) == ["Kundenname", "Defizite", "Lösung & Nutzen"], (
        f"Unexpected header: {all_rows[0]}"
    )

    # Exactly 2 data rows (one per company)
    data_rows = all_rows[1:]
    assert len(data_rows) == 2, f"Expected 2 data rows, got {len(data_rows)}"

    # Each data row has 3 columns (Kundenname must not be empty)
    for i, row in enumerate(data_rows):
        assert len(row) == 3, f"Row {i} has {len(row)} columns, expected 3"
        assert row[0], f"Row {i} Kundenname is empty"


# ---------- test_second_sheet_same_sorted_order (DIFF-04) ----------

def test_second_sheet_same_sorted_order(tmp_path):
    """Sheet 2 rows appear in the SAME sorted order as the 'Leads' sheet body."""
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    headers = ["Kundenname", "Website"]
    rows = [
        ["Firma A", "https://a.ch"],
        ["Firma B", "https://b.ch"],
    ]
    _make_xlsx(inp, headers, rows)

    run(Config(input=str(inp), output=str(out)))

    wb = openpyxl.load_workbook(out)
    leads_ws = wb["Leads"]
    arg_ws = wb["Verkaufsargumente"]

    # Get Kundenname column from both sheets (skip header row)
    leads_names = [r[0] for r in leads_ws.iter_rows(min_row=2, values_only=True)]
    arg_names = [r[0] for r in arg_ws.iter_rows(min_row=2, values_only=True)]

    assert leads_names == arg_names, (
        f"Order mismatch: Leads={leads_names}, Argumente={arg_names}"
    )


# ---------- test_leads_sheet_unchanged (regression guard) ----------

def test_leads_sheet_unchanged(tmp_path):
    """The 'Leads' sheet header and row count are unchanged by adding the second sheet."""
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    headers = ["Kundennummer", "Kundenname", "Branche", "Ort", "Website"]
    rows = [
        [1, "Firma A", "Maler", "Zürich", "https://a.ch"],
        [2, "Firma B", "Garage", "Bern", "https://b.ch"],
    ]
    _make_xlsx(inp, headers, rows)

    run(Config(input=str(inp), output=str(out)))

    wb = openpyxl.load_workbook(out)
    assert "Leads" in wb.sheetnames

    leads_ws = wb["Leads"]
    all_rows = list(leads_ws.iter_rows(values_only=True))

    # Header row: original 5 + 2 score columns + 1 reason column = 8
    assert list(all_rows[0][:5]) == headers
    assert all_rows[0][5] == "Website-Bedarf (1-5)"
    assert all_rows[0][6] == "Zahlungskräftigkeit (1-5)"

    # Row count: 1 header + 2 data rows
    assert len(all_rows) == 3, f"Expected 3 rows (1 header + 2 data), got {len(all_rows)}"


# ---------- test_csv_companion_argumente (DIFF-04) ----------

def test_csv_companion_argumente(tmp_path):
    """When write_csv=True, a companion *_argumente.csv is written next to the main CSV."""
    inp = tmp_path / "in.xlsx"
    # Use xlsx input but request csv output via write_csv=True
    out = tmp_path / "out.xlsx"
    headers = ["Kundenname", "Website"]
    rows = [
        ["Firma A", "https://a.ch"],
    ]
    _make_xlsx(inp, headers, rows)

    run(Config(input=str(inp), output=str(out), write_csv=True))

    companion = tmp_path / "out_argumente.csv"
    assert companion.exists(), f"Companion argumente.csv not found at {companion}"

    # Check that it has the correct headers
    text = companion.read_text(encoding="utf-8-sig")
    reader = csv.reader(text.splitlines())
    header_row = next(reader)
    assert header_row == ["Kundenname", "Defizite", "Lösung & Nutzen"], (
        f"Unexpected CSV header: {header_row}"
    )


def test_csv_output_companion_argumente(tmp_path):
    """When output is a .csv file, a companion *_argumente.csv is written."""
    inp = tmp_path / "in.csv"
    out = tmp_path / "out.csv"
    inp.write_text("Kundenname,Website\nFirma A,https://a.ch\n", encoding="utf-8")

    run(Config(input=str(inp), output=str(out)))

    companion = tmp_path / "out_argumente.csv"
    assert companion.exists(), f"Companion argumente.csv not found at {companion}"

    text = companion.read_text(encoding="utf-8-sig")
    reader = csv.reader(text.splitlines())
    header_row = next(reader)
    assert header_row == ["Kundenname", "Defizite", "Lösung & Nutzen"]
