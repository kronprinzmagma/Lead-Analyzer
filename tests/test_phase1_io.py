"""Phase-1-Tests: Tabellen-I/O, URL-Spalten-Erkennung, verlustfreier Sort.

Decken die Phase-1-Erfolgskriterien ab (IO-01..07): Original-Spalten unverändert,
genau zwei ganzzahlige Score-Spalten, tolerante URL-Erkennung mit klarer Fehlermeldung,
absteigender Sort ohne Zeilenverlust.
"""

from __future__ import annotations

import openpyxl
import pytest

from lead_analyzer import scoring, table_io
from lead_analyzer.config import Config
from lead_analyzer.models import RowRecord, RowResult
from lead_analyzer.pipeline import run
from lead_analyzer.table_io import InputError


def _make_xlsx(path, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(path)


# ---------- URL-Spalten-Erkennung (IO-02) ----------

@pytest.mark.parametrize("name", ["URL", "Website", "Webseite", "Web", "homepage", "Website URL", "Web-Adresse"])
def test_detect_url_column_variants(name):
    assert table_io.detect_url_column(["Kundenname", name, "Ort"]) == name


def test_detect_url_column_missing_raises():
    with pytest.raises(InputError):
        table_io.detect_url_column(["Kundenname", "Ort", "Branche"])


# ---------- Sort: absteigend + verlustfrei (IO-05) ----------

def test_stable_sort_desc_and_lossless():
    recs = [RowRecord(i, {"x": i}) for i in range(5)]
    results = [
        RowResult(0, bedarf=3, zahl=2),
        RowResult(1, bedarf=5, zahl=1),
        RowResult(2, bedarf=3, zahl=4),
        RowResult(3, bedarf=5, zahl=5),
        RowResult(4, bedarf=1, zahl=1),
    ]
    ordered = scoring.stable_sort(list(zip(recs, results)))
    assert len(ordered) == len(recs)  # nichts verloren
    order_idx = [rec.index for rec, _ in ordered]
    # Erwartung: (5,5)->3, (5,1)->1, (3,4)->2, (3,2)->0, (1,1)->4
    assert order_idx == [3, 1, 2, 0, 4]


def test_stable_sort_tiebreak_keeps_original_order():
    recs = [RowRecord(i, {}) for i in range(3)]
    results = [RowResult(i, bedarf=4, zahl=3) for i in range(3)]  # alle gleich
    ordered = scoring.stable_sort(list(zip(recs, results)))
    assert [rec.index for rec, _ in ordered] == [0, 1, 2]


def test_clamp_score_bounds():
    assert scoring.clamp_score(0) == 1
    assert scoring.clamp_score(9) == 5
    assert scoring.clamp_score(3) == 3


# ---------- End-to-End: xlsx rein -> xlsx raus (IO-01,03,04,06,07) ----------

def test_end_to_end_xlsx_preserves_columns_and_appends_scores(tmp_path):
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    headers = ["Kundennummer", "Kundenname", "Branche", "Ort", "Website"]
    rows = [
        [1, "Firma A", "Maler", "Zürich", "https://a.ch"],
        [2, "Firma B", "Garage", "Bern", "https://b.ch"],
    ]
    _make_xlsx(inp, headers, rows)

    summary = run(Config(input=str(inp), output=str(out)))
    assert summary["url_column"] == "Website"
    assert summary["rows_processed"] == 2

    wb = openpyxl.load_workbook(out)
    got = list(wb.active.iter_rows(values_only=True))
    out_headers = list(got[0])
    # Original-Header unverändert + Reihenfolge
    assert out_headers[:5] == headers
    assert out_headers[5:7] == ["Website-Bedarf (1-5)", "Zahlungskräftigkeit (1-5)"]
    # Score-Werte ganzzahlig 1..5, nie leer
    for r in got[1:]:
        b, z = r[5], r[6]
        assert isinstance(b, int) and 1 <= b <= 5
        assert isinstance(z, int) and 1 <= z <= 5
    # Original-Zellwerte unverändert durchgereicht
    body = sorted(got[1:], key=lambda r: r[0])
    assert body[0][:5] == (1, "Firma A", "Maler", "Zürich", "https://a.ch")


def test_limit_only_processes_first_n(tmp_path):
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    headers = ["Name", "Web"]
    rows = [[f"F{i}", f"https://f{i}.ch"] for i in range(10)]
    _make_xlsx(inp, headers, rows)
    summary = run(Config(input=str(inp), output=str(out), limit=3))
    assert summary["rows_processed"] == 3
    wb = openpyxl.load_workbook(out)
    assert len(list(wb.active.iter_rows())) - 1 == 3


def test_csv_roundtrip(tmp_path):
    inp = tmp_path / "in.csv"
    out = tmp_path / "out.csv"
    inp.write_text("Kundenname,Website\nFirma A,https://a.ch\nFirma B,https://b.ch\n", encoding="utf-8")
    run(Config(input=str(inp), output=str(out)))
    text = out.read_text(encoding="utf-8-sig")
    assert "Website-Bedarf (1-5)" in text
    assert "Firma A" in text


def test_missing_url_column_raises_clear_error(tmp_path):
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    _make_xlsx(inp, ["Kundenname", "Ort"], [["A", "Zürich"]])
    with pytest.raises(InputError):
        run(Config(input=str(inp), output=str(out)))
