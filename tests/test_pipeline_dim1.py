"""Phase-2-Tests: Dimension-1-Verdrahtung im Pipeline-Layer (offline).

Deckt ROB-01 (leere/kaputte URL -> Bedarf 5, leere URL OHNE Netz), ROB-03
(Per-Row-Exception-Boundary isoliert eine kaputte Zeile; der Lauf geht weiter)
und die dead->5-Override-Logik ab. Plus eine Offline-Integration über die echte
`data/sample_input.xlsx`: `fetch.fetch` wird durch einen deterministischen Fake
ersetzt — KEIN Test fasst das Netz an (autouse-Sperre aus conftest bleibt aktiv).

Stil angelehnt an `test_phase1_io.py` (openpyxl-Helfer, Docstrings).
"""

from __future__ import annotations

import os

import openpyxl
import pytest

from lead_analyzer import fetch, scoring
from lead_analyzer.analyzers import existence, payment
from lead_analyzer.config import Config
from lead_analyzer.models import RowRecord, RowResult
from lead_analyzer.pipeline import analyze_row, run

from conftest import make_fetch_result

_CFG = Config(input="x", output="y")
_URL_COL = "Website"


def _rec(url):
    return RowRecord(index=0, cells={_URL_COL: url})


# --------------------------------------------------------------------------- #
# scoring.bedarf_from_dim1 — provisorische Abbildung                           #
# --------------------------------------------------------------------------- #

def test_bedarf_from_dim1_dead_causes_are_5():
    from lead_analyzer.models import DimensionVerdict
    # Override hängt am dead-Flag, NICHT am Reason-Text (Review L5).
    for reason in ("nicht erreichbar (Timeout)", "geparkt/Platzhalter"):
        v = DimensionVerdict(1, "severe", reason, dead=True)
        assert scoring.bedarf_from_dim1(v) == 5


def test_bedarf_from_dim1_text_alone_does_not_force_5():
    """Eine severe-Verdict mit 'tot klingendem' Text aber dead=False darf NICHT 5 sein."""
    from lead_analyzer.models import DimensionVerdict
    v = DimensionVerdict(1, "severe", "nicht erreichbar (Timeout)", dead=False)
    assert scoring.bedarf_from_dim1(v) == 4  # severe-nicht-tot


def test_bedarf_from_dim1_social_severe_is_4():
    from lead_analyzer.models import DimensionVerdict
    v = DimensionVerdict(1, "severe", "Social-only")
    assert scoring.bedarf_from_dim1(v) == 4


def test_bedarf_from_dim1_gap_and_ok_are_3():
    from lead_analyzer.models import DimensionVerdict
    assert scoring.bedarf_from_dim1(DimensionVerdict(1, "gap", "dünner Inhalt")) == 3
    assert scoring.bedarf_from_dim1(DimensionVerdict(1, "ok", "erreichbar")) == 3


# --------------------------------------------------------------------------- #
# analyze_row — ROB-01 / dead->5 / 403-neutral / zahl-Platzhalter             #
# --------------------------------------------------------------------------- #

def test_empty_url_no_network(monkeypatch):
    """Leere URL -> Bedarf 5 'keine Website', OHNE fetch.fetch-Aufruf."""
    called = {"n": 0}

    def spy(*a, **k):
        called["n"] += 1
        raise AssertionError("fetch darf bei leerer URL NICHT aufgerufen werden")

    monkeypatch.setattr(fetch, "fetch", spy)
    res = analyze_row(_rec(None), _URL_COL, _CFG)
    assert res.bedarf == 5
    assert "keine Website" in res.reason
    assert called["n"] == 0
    # zahl = echte konservative Schätzung (kein Name/Branche, fr/soup None -> 2)
    assert res.zahl == payment.estimate(_rec(None), None, None, _CFG).zahl


def test_empty_string_url_no_network(monkeypatch):
    def spy(*a, **k):
        raise AssertionError("kein fetch bei leerem String")

    monkeypatch.setattr(fetch, "fetch", spy)
    res = analyze_row(_rec("   "), _URL_COL, _CFG)
    assert res.bedarf == 5
    assert "keine Website" in res.reason


def test_dead_unreachable_is_bedarf_5(monkeypatch):
    monkeypatch.setattr(
        fetch, "fetch",
        lambda c, cfg: make_fetch_result(ok=False, status=None, html=None,
                                         error="nicht erreichbar"),
    )
    res = analyze_row(_rec("htp://naehatelier-sutter"), _URL_COL, _CFG)
    assert res.bedarf == 5
    # res.reason kommt jetzt aus reasons.build ("Dim1 … (nicht erreichbar …) →
    # Bedarf 5") -> Substring statt startswith (reconciled in 03-04).
    assert "nicht erreichbar" in res.reason
    assert res.verdicts and res.verdicts[0].dim == 1


def test_parked_is_bedarf_5(monkeypatch):
    monkeypatch.setattr(
        fetch, "fetch",
        lambda c, cfg: make_fetch_result(
            html="<html><title>This domain is for sale</title><body>buy this domain</body></html>"),
    )
    res = analyze_row(_rec("https://parked.ch"), _URL_COL, _CFG)
    assert res.bedarf == 5


def test_block_403_is_neutral_not_5(monkeypatch):
    # Clean-https-WAF-Block: Dim1 gap + Dim2 ok + Dim3 gap + Dim4/5/6 neutral
    # -> G=2 -> Band 3. !=5 ist die load-bearing Invariante; ==3 ist der dokumentierte
    # 6-Dim-Wert — identisch in test_pipeline_bedarf.py.
    # BED-03: ohne Body (html=None) gibt es KEIN viewport-meta (soup None) -> die
    # Dim-3-Heuristik liefert korrekt 'gap' statt des alten Platzhalter-'ok'
    # (Mobil-Tauglichkeit eines geblockten Bodys ist nicht nachweisbar). Bänder unverändert.
    monkeypatch.setattr(
        fetch, "fetch",
        lambda c, cfg: make_fetch_result(
            ok=False, status=403, html=None,
            url="https://waf.ch/", final_url="https://waf.ch/", ssl_ok=True),
    )
    res = analyze_row(_rec("https://waf.ch"), _URL_COL, _CFG)
    assert res.bedarf != 5
    assert res.bedarf == 3
    assert "blockiert" in res.reason


def test_social_only_high_but_not_5(monkeypatch):
    # BED-03: dieser Test prüft die Dim-1-Logik (Social-only -> severe, aber NICHT
    # dead-5). Das Fixture bekommt ein viewport-meta, damit der ab Phase 6 reale
    # Dim-3-Befund 'ok' ist und KEINE Dim-3-Lücke das Dim-1-Verhalten überlagert
    # (sonst kippte die G-Summe der ohnehin dünnen Social-Seite auf Band 5). Damit
    # bleibt der dokumentierte Wert 4 erhalten; Scoring-Bänder unverändert.
    monkeypatch.setattr(
        fetch, "fetch",
        lambda c, cfg: make_fetch_result(
            final_url="https://facebook.com/firma",
            html='<html><head><meta name="viewport" content="width=device-width">'
                 "</head><body>Profil</body></html>"),
    )
    res = analyze_row(_rec("https://facebook.com/firma"), _URL_COL, _CFG)
    assert res.bedarf == 4  # severe-not-dead


def test_reachable_thin_is_4(monkeypatch):
    monkeypatch.setattr(
        fetch, "fetch",
        lambda c, cfg: make_fetch_result(html="<html><body>kurz</body></html>"),
    )
    res = analyze_row(_rec("https://thin.ch"), _URL_COL, _CFG)
    # 6-Dim reconciled (war provisorisch 3 unter reinem Dim-1-Pfad): dünner Inhalt
    # (Dim1 gap) + fehlender Title/Meta/H1 (Dim4 gap) + kein Markup (Dim5 severe)
    # + fehlender Kontakt/Impressum (Dim6 gap) -> Bedarf 4.
    assert res.bedarf == 4


def test_zahl_is_real_estimate(monkeypatch):
    monkeypatch.setattr(fetch, "fetch", lambda c, cfg: make_fetch_result())
    res = analyze_row(_rec("https://ok.ch"), _URL_COL, _CFG)
    # link-lose Default-Seite, kein Name/Branche -> Group C leer -> konservative 2
    assert res.zahl == payment.estimate(_rec("https://ok.ch"), None, None, _CFG).zahl


# --------------------------------------------------------------------------- #
# ROB-03 — Per-Row-Boundary: eine kaputte Zeile killt den Lauf nicht          #
# --------------------------------------------------------------------------- #

def test_row_boundary_degrades_not_raises(monkeypatch):
    """existence.analyze wirft mitten in der Zeile -> degradiertes RowResult, kein Crash."""
    def boom(fr):
        raise ValueError("kaputt")

    monkeypatch.setattr(fetch, "fetch", lambda c, cfg: make_fetch_result())
    monkeypatch.setattr(existence, "analyze", boom)
    res = analyze_row(_rec("https://boom.ch"), _URL_COL, _CFG)  # darf NICHT raisen
    assert res.bedarf == 5
    assert res.reason.startswith("Fehler:")


def test_run_continues_after_bad_row(monkeypatch, tmp_path):
    """Eine kaputte Zeile bricht run() nicht ab — alle Zeilen kommen raus."""
    inp = tmp_path / "in.xlsx"
    out = tmp_path / "out.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Kundenname", "Website"])
    ws.append(["Gut", "https://gut.ch"])
    ws.append(["Kaputt", "https://kaputt.ch"])
    wb.save(inp)

    def maybe_boom(c, cfg):
        if any("kaputt" in u for u in c):
            raise RuntimeError("netz-explosion")
        return make_fetch_result()

    monkeypatch.setattr(fetch, "fetch", maybe_boom)
    summary = run(Config(input=str(inp), output=str(out)))
    assert summary["rows_processed"] == 2
    wb2 = openpyxl.load_workbook(out)
    assert len(list(wb2.active.iter_rows())) - 1 == 2  # beide Zeilen geschrieben


# --------------------------------------------------------------------------- #
# Offline-Integration über die echte Sample-Datei (ROB-01/03, AC10)           #
# --------------------------------------------------------------------------- #

_SAMPLE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "data", "sample_input.xlsx",
)


@pytest.mark.skipif(not os.path.exists(_SAMPLE), reason="sample_input.xlsx fehlt")
def test_sample_offline(monkeypatch, tmp_path):
    """Voller Lauf über data/sample_input.xlsx mit gemocktem fetch — KEIN Netz.

    Deterministischer Fake: erreichbare Nicht-Leer-URLs liefern Inhalt; der
    bewusst kaputte 'naehatelier-sutter'-Host (kein DNS) liefert 'nicht erreichbar'.
    Erwartung: len(out)==len(in), jedes bedarf int in [1,5], Kiosk (leere URL) und
    Nähatelier (kaputte URL) beide == 5.
    """
    out = tmp_path / "out.xlsx"

    body = "<html><head><title>Firma</title></head><body>" + ("Wort " * 400) + "</body></html>"

    def fake_fetch(candidates, cfg):
        # 'naehatelier-sutter' (kein Punkt -> DNS-Fail) simuliert unerreichbar.
        if any("naehatelier-sutter" in u for u in candidates):
            return make_fetch_result(ok=False, status=None, html=None,
                                     error="nicht erreichbar")
        return make_fetch_result(html=body)

    monkeypatch.setattr(fetch, "fetch", fake_fetch)

    summary = run(Config(input=_SAMPLE, output=str(out)))
    n_in = summary["rows_processed"]

    wb = openpyxl.load_workbook(out)
    rows = list(wb.active.iter_rows(values_only=True))
    header = list(rows[0])
    body_rows = rows[1:]

    assert len(body_rows) == n_in  # len(out) == len(in)

    b_idx = header.index("Website-Bedarf (1-5)")
    name_idx = header.index("Kundenname")
    web_idx = header.index("Website")

    for r in body_rows:
        b = r[b_idx]
        assert isinstance(b, int) and 1 <= b <= 5

    # Kiosk (leere URL) und Nähatelier (kaputte URL) -> beide Bedarf 5.
    by_name = {r[name_idx]: r for r in body_rows}
    kiosk = next((r for n, r in by_name.items() if n and "Kiosk" in str(n)), None)
    naeh = next(
        (r for r in body_rows
         if r[web_idx] and "naehatelier-sutter" in str(r[web_idx])),
        None,
    )
    assert kiosk is not None and kiosk[b_idx] == 5
    assert naeh is not None and naeh[b_idx] == 5
